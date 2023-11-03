#!/usr/bin/python3

from openpyxl import load_workbook

#학점 구하기
def grade(total, totallist, num):
    #학점을 받을 수 있는 학생 수
    A = int(0.3 * num)
    B = int(0.7 * num)
    A_Plus = min(A, int(0.5 * A))
    B_Plus = min(B, int(0.5 * B))
    C_Plus = min(int(0.5 * (0.9 * num)))
    

    #(totallist)학점을 받을 수 있는 마지막 학생의 점수와 각 행의 total의 점수를 비교해서 학점 넣기
    if total < 40:
        return 'F'
    elif total >= totallist[int(0.9 * num) -1]:
        return 'C'
    elif total >= totallist[C_Plus - 1]:
        return 'C+'
    elif  total >= totallist[B - 1]:
        return 'B'
    elif total >= totallist[B_Plus - 1]:
        return 'B+'
    elif total >= totallist[A - 1]:
        return 'A'
    else:
        return 'A+'
     
wb = load_workbook(filename='student.xlsx')
ws = wb.active

totallist = []

# 두번째 행부터 total구하기
for row in ws.iter_rows(min_row=2):
    total = ((row[2].value*0.3) + (row[3].value*0.35) + (row[4].value*0.34) + (row[5].value*0.1))+0.9
    row[6].value = total

#total기준으로 정렬한 리스트 생성
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    total = row[6].value
    totallist.append(total)

totallist = sorted(totallist)

num = ws.max_row - 1

#ws순회하면서 학점 넣기
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    row[7].value = grade(row[6].value, totallist, num)

wb.save(filename='student.xlsx')